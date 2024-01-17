from __future__ import print_function
from traceback import print_tb
import requests
import os.path
from ImagePDF import ImagePDF
from ReadQR import ReadQR 

try:
    from bs4 import BeautifulSoup
except ImportError:
    from BeautifulSoup import BeautifulSoup
import urllib3
import xlsxwriter
from openpyxl import load_workbook

requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS = 'ALL:@SECLEVEL=1'
class ScrapSAT():
    url = " "
    # Local variable corresponding to identificaction data for a user
    datosDeIdentificacion = []
    # Local variable for the ubication data for a user
    datosDeUbicacion = []
    # Fiscal characteristics of a user
    caracteristicasFiscales = []
    #identification data Complete
    datosIdentificacionCom = []
    #Name of the columns is identification data
    datosIdentificacionColum = []
    #Ubicacion data complete
    datosUbicaCom = []
    #Name of the columns in ubication data
    datosUbicaColumn = []
    #Fiscal Characteristics Complete
    caracterCom = []
    #Name of the columns in fiscal characteristics data
    caracterColum = []
    def __init__(self,URL,datosDeIdentificacion,datosDeUbicacion,caracteristicasFiscales,datosIdentificacionCom,datosIdentificacionColum,datosUbicaCom,datosUbicaColumn,caracterCom,caracterColum):
        self.url= URL
        self.datosDeIdentificacion = datosDeIdentificacion
        self.datosDeUbicacion = datosDeUbicacion
        self.caracteristicasFiscales = caracteristicasFiscales
        self.datosIdentificacionCom =datosIdentificacionCom
        self.datosIdentificacionColum = datosIdentificacionColum
        self.datosUbicaCom = datosUbicaCom
        self.datosUbicaColumn = datosUbicaColumn
        self.caracterCom = caracterCom
        self.caracterColum = caracterColum
        # Given a certain URL for SAT data user, fills the local lists of the user information
        #URL the URL to use
    def generaListas(self, url):
            page = requests.get(url)
            soup = BeautifulSoup(page.content, "html.parser")
            results = soup.find_all('td')
            i = 0
            position = 0
            information ={}
            for resultado in results:
                # Si estamos en datos de identificacion, apartir de la posicion 1 metemos a la lista.
                if(position ==0):
                    if(i>0 and i<13):
                        tagAString = resultado.contents
                        self.datosDeIdentificacion.append(tagAString)
                    if(i == 13):
                        i = 0
                        position = 1
                    else: i = i+1
                if(position == 1):
                    if(i>3 and i<24):
                        tagAString = resultado.contents
                        self.datosDeUbicacion.append(tagAString)
                    if(i==24):
                        i=0
                        position = 2
                    else: i=i+1
                if(position==2):
                    if(i>3 and i<8):
                        tagAString = resultado.contents
                        self.caracteristicasFiscales.append(tagAString)
                    if(i==9):
                        i=0
                        position = 2
                    else: i=i+1
    '''
    Split list of values into 2 lists, one containing colunm values and other with the values of the colunms 
    '''
    def parteListas(self, lista,lista2,lista3):
        i = 0
        while(i<len(lista)):
            if(i%2 != 0):
                lista2.append(str(lista[i]))
            else: lista3.append(str(lista[i]))
            i+=1
            

    '''Given a String in format [<span style="font-weight: bold;">Entidad Federativa:</span>]
    The method returns only the information between tags'''
    def parteData(self, lista1):
        i = 0
        partidoCompletoColum = []
        for item in lista1:
            partido = item.split(";")
            partidopal = partido[1].split(":")
            partidoCompleto = partidopal[0].split(">")
            partidoCompletoColum.append(partidoCompleto[1])
        return partidoCompletoColum
    '''
    Given a list of strings with certain format, splits the string into the format we want it to be
    '''
    def parteData2(self, lista1):
        i = 0
        partidoCompleto = []
        for item in lista1:
            partido = item.split("'")
            if(len(partido)>1):
                partidoCompleto.append(partido[1])
            else: partidoCompleto.append(" ")

        return partidoCompleto
    '''
    Writes the strings of the lists given in an excel as columns
    '''
    def escribeColumnas(self, listaIden, listaUbi,listaCar, worksheet):
        row = 0
        column = 0
        for item in listaIden :
            worksheet.write(row, column, item)
            column += 1
        for item in listaUbi :
            worksheet.write(row, column, item)
            column += 1
        for item in listaCar :
            worksheet.write(row, column, item)
            column += 1
    '''
    Writes the strings of the lists given in an excel (each one corresponding it´s column)
    Preguntar si el elemento ya se encuentra en la tabla, si sí, entonces no escribas
    '''
    def escribeColumnasData(self, listaIden, listaUbi,listaCar, row, worksheet):
        column = 0
        for item in listaIden :
            worksheet.write(row, column, item)
            column += 1
        for item in listaUbi :
            worksheet.write(row, column, item)
            column += 1
        for item in listaCar :
            worksheet.write(row, column, item)
            column += 1
    '''
    Saves the information from an excel to a list.
    @param cadena - The name of the file
    @return a list with the whole data from an excel
    '''
    def leeExcel(self,cadena):
        listaComp = []
        wb = load_workbook(cadena)
        sheet = wb.active
        for row in sheet.values:
            for value in row:
                if(str(value)== "None"):
                    continue
                else:
                    listaComp.append(str(value))
        return listaComp
    
    def escribeConWritexl(self,row_count, worksheet,lista):
        colum = 0
        row = 0
        total=0
        print("Datos a escribir:", lista)
        while(row<=row_count) :
            if(total % 18 == 0 and total>0):
                row+=1
                if(row==row_count):
                    return
                colum = 0
                worksheet.write(row, colum,lista[total])
                colum+=1
            else:
                worksheet.write(row, colum,lista[total])
                colum +=1
           
            total+=1   


    def encuentraEnLista(self, item, lista2):
        i=0
        while(i < len(lista2)):
            print("Comparando",lista2[i]," con",item)
            if (lista2[i] == item):
                return True
            else:
                i = i+1
        return False

'''
Main method, runs the complete algorithm as it follows:
    1- Reads all the PDF´s files in this same file, and gives the QR code that the PDF contains
    2- Reads the QR Code images, and gives a list with the URL´s of the QR codes.
    3- Given the URL codes, reads each one and saves data into 2 lists, columns and column values.
    4- Splits column values and returns information in a specific value
    5- Writes data in an excel named datos.xlsx
'''
def main():
    existe=False
    scrap1 = ScrapSAT("1",[],[],[],[],[],[],[],[],[])
    # Image names 
    nombreImagenes = []
    workdir = os.getcwd()
    imagenees = ImagePDF(nombreImagenes)
    nombreImagenes = imagenees.getQRSFromPDF(workdir,nombreImagenes)
    nombreImagenes =[]
    for each_path in  os.listdir(workdir):
        if(".png") in each_path:
            nombreImagenes.append(each_path)
    #lista de URLS
    listaURL = []
    qrs = ReadQR(listaURL)
    listaURL = qrs.readQr(nombreImagenes)
    print("URLS a leer:", listaURL)
    entrada = 0
    if(os.path.exists("datos.xlsx")):
        existe= True
        listacompletaExcelLEido = scrap1.leeExcel("datos.xlsx")
        wb = load_workbook("datos.xlsx", enumerate)
        sheet = wb.active
        #Número máximo de filas
        row_count = sheet.max_row
    workbook = xlsxwriter.Workbook('datos.xlsx')
    worksheet = workbook.add_worksheet() 
    if(existe):
        scrap1.escribeConWritexl(row_count, worksheet, listacompletaExcelLEido)   
    while(entrada< len(listaURL)+1):
        #Local variable corresponding to identificaction data for a user
        datosDeIdentificacion = []
        # Local variable for the ubication data for a user
        datosDeUbicacion = []
        # Fiscal characteristics of a user
        caracteristicasFiscales = []
        #identification data Complete
        datosIdentificacionCom = []
        #Name of the columns is identification data
        datosIdentificacionColum = []
        #Ubicacion data complete
        datosUbicaCom = []
        #Name of the columns in ubication data
        datosUbicaColumn = []
        #Fiscal Characteristics Complete
        caracterCom = []
        #Name of the columns in fiscal characteristics data
        caracterColum = []
        if(entrada>0):
            URLPrueba = listaURL[entrada-1]
        else: URLPrueba = listaURL[entrada]
        scrap = ScrapSAT(URLPrueba,datosDeIdentificacion,datosDeUbicacion,caracteristicasFiscales,datosIdentificacionCom,datosIdentificacionColum,datosUbicaCom,datosUbicaColumn,caracterCom,caracterCom)
        scrap.generaListas(URLPrueba)
        scrap.parteListas(datosDeIdentificacion,datosIdentificacionCom,datosIdentificacionColum)
        scrap.parteListas(datosDeUbicacion, datosUbicaCom, datosUbicaColumn)
        scrap.parteListas(caracteristicasFiscales,caracterCom, caracterColum)
        # These are the lists with complete information
        datosIdentificacionColumnas = scrap.parteData(datosIdentificacionColum)
        datosUbicacionColumnas = scrap.parteData(datosUbicaColumn)
        datosCaracterisiticasColumnas = scrap.parteData(caracterColum)
        datosIdenComplete = scrap.parteData2(datosIdentificacionCom)
        datosUbiComplete = scrap.parteData2(datosUbicaCom)
        datosCaracComplete = scrap.parteData2(caracterCom)
        #Empieza a escribir
        if(entrada == 0) :
            scrap.escribeColumnas(datosIdentificacionColumnas,datosUbicacionColumnas,datosCaracterisiticasColumnas, worksheet)
        else: 
            if(existe):
                scrap.escribeColumnasData(datosIdenComplete, datosUbiComplete, datosCaracComplete, row_count, worksheet)
                row_count+=1
            else:
                scrap.escribeColumnasData(datosIdenComplete, datosUbiComplete, datosCaracComplete, entrada, worksheet)

        entrada +=1
    workbook.close()

if __name__ == '__main__':
        main()